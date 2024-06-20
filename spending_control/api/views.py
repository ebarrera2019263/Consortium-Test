
import io

from django.http import HttpResponse
from django.db.models import Sum
from django.conf import settings
from django.db import transaction

from openpyxl import load_workbook
from openpyxl.styles import Protection

from decimal import Decimal

from pathlib import Path

from rest_framework.generics import ListCreateAPIView, RetrieveUpdateAPIView
from rest_framework.response import Response
from rest_framework import status

from spending_control.models import Spending
from spending_control.api.serializers import SpendingSerializer


class SpendingControlListCreateAPIView(ListCreateAPIView):
    queryset = Spending.objects.all()
    serializer_class = SpendingSerializer


class SpendingRetrieveAPIView(RetrieveUpdateAPIView):
    queryset = Spending.objects.all()
    serializer_class = SpendingSerializer

    def partial_update(self, request, *args, **kwargs):
        
        return Response({'message': 'OK'}, status=status.HTTP_206_PARTIAL_CONTENT)


def generate_liquidation_certificate(request, pk):
    instance = Spending.objects.get(pk=pk)
    liquidations = instance.liquidations.all()

    file_path = rf'{settings.BASE_DIR}\spending_control\templates\spending_control\liquidation_certificate.xlsx'
    file_path = Path(file_path)
    
    buffer = io.BytesIO()
    
    wb = load_workbook(file_path)
    ws = wb['CONOCIMIENTO']
    
    ws.protection.sheet = True
    ws.protection.password = 'Archivo de prueba'
    
    unlock_cells = ws['A1:S50']
    
    for item in unlock_cells:
        for cell in item:
            if cell.row != 20:
                cell.protection = Protection(locked=False)
                
                
    detail = 'A continuación se adjuntan los siguientes comprobantes:\n\n'
    
    for liquidation in liquidations:
        detail += f'{liquidation.entity} {liquidation.invoice_number} {liquidation.total_value}\n'
        
    ws['B20'] = detail
    
    wb.save(buffer)
    buffer.seek(0)
    
    instance.liquidation_certificate.save('Conocimiento_Prueba.xlsx', buffer)
    instance.save()
    
    response = HttpResponse(content_type='application/ms-excel')
    contenido = 'attachment; filename={}'.format(
        f'Conocimiento_Prueba.xlsx')
    response['Content-Disposition'] = contenido
    wb.save(response)
    
    return response